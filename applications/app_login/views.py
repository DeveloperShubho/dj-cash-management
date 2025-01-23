
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from django.shortcuts import render, redirect, get_object_or_404
from django.core.exceptions import ObjectDoesNotExist
from django.views.decorators.cache import never_cache
from django.db.models.functions import Cast
from django.db.models import FloatField
from django.utils.timezone import now
from django.http import HttpResponse
from applications.constants import *
from django.contrib import messages
from django.db.models import Sum
from django.db.models import Q
from dotenv import load_dotenv
from openpyxl import Workbook
from io import BytesIO
from .models import *
from .forms import *
import os




# Load environment variables from the .env file
load_dotenv()

##########################################################################################################################################

# Function to check verification key
@never_cache
def verification_view(request):
    if 'is_verification' in request.session.keys():
        return redirect('/login/')
    
    if request.method == 'POST':
        input_key = request.POST.get('key')
        correct_key = os.getenv('VERIFICATION_KEY', '')

        if input_key == correct_key:
            request.session['is_verification'] = True
            return redirect('/login/')
        else:
            messages.error(request, "Invalid Verification Key. Please try again.")
    
    return render(request, 'index.html')

##########################################################################################################################################

@never_cache
def login_view(request):
    if 'is_login' in request.session.keys():
        return redirect('dashboard')
    
    if 'is_verification' in request.session.keys():  
        if request.method == 'POST':
            input_username = request.POST.get('username')
            input_password = request.POST.get('password')

            # Retrieve the user(s) from the database based on the loginname
            users = Pcash100User.objects.filter(loginname_100=input_username)

            if not users:
                messages.error(request, "Invalid Username. Please try again.")
                return redirect('login')

            # Iterate through all users with the given loginname and check the password
            for user in users:
                md5_hash = user.loginpwd_100
                if check_md5_password(md5_hash, input_password):
                    # Set session data for the logged-in user
                    request.session['is_login'] = True
                    request.session['current_user'] = input_username
                    request.session['executive_user_id'] = user.id_100
                    return redirect('dashboard')  
                
            # If no valid user is found with the correct password
            messages.error(request, "Invalid Password. Please try again.")
            return redirect('login')
        
        return render(request, 'login.html')   
    
    return redirect('/')

##########################################################################################################################################

@never_cache
def logout_view(request):
    if 'is_login' in request.session.keys():
        request.session.flush()
        return redirect('/')
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def credit_view(request):
    if 'is_login' in request.session.keys():

        request.session['current_page']='credit_view'

        if request.method == 'POST':
            form = CreditForm(request.POST)
            if form.is_valid():
                credit_type = request.POST['credit_type']

                if credit_type == 'internal_credit':
                    # Retrieving form data
                    cash_withdrawal_date = format_date(form.cleaned_data.get('withdrawal_date'))  
                    bank_id = form.cleaned_data.get('bank')  
                    instrument_type = form.cleaned_data.get('instrument_type') 
                    cheque_number = form.cleaned_data.get('cheque_number')
                    instrument_date = format_date(form.cleaned_data.get('instrument_date'))  
                    narration = form.cleaned_data.get('narration_1')  
                    amount = form.cleaned_data.get('amount_1')  
                    withdrawal_by = form.cleaned_data.get('withdrawal_by') 
                    

                    
                    payment_date = ''
                    paid_by = '' 
                    collected_by = ''   
                        

                if credit_type == 'other_payment':
                    # Retrieving form data 
                    payment_date = format_date(form.cleaned_data.get('payment_date'))  
                    narration = form.cleaned_data.get('narration_2')  
                    amount = form.cleaned_data.get('amount_2') 
                    paid_by = form.cleaned_data.get('paid_by')  
                    collected_by = form.cleaned_data.get('collected_by') 


                    cash_withdrawal_date = get_current_date_ymd() 
                    bank_id = 0
                    instrument_type = ''
                    cheque_number = ''
                    instrument_date = '' 
                    withdrawal_by = ''


                # Creating an instance of Pcash250Addcredit and saving the data
                Pcash250Addcredit.objects.create(
                    type_250=credit_type,
                    cash_withdrawal_ymd_250=cash_withdrawal_date,
                    bank_id_250=bank_id,
                    instrument_type_250=instrument_type,
                    instrument_no_250=cheque_number,
                    instrument_ymd_250=instrument_date,
                    narration_250=narration,
                    paid_by_250=paid_by,
                    amount_250=amount,
                    withdrawal_by_250=withdrawal_by,
                    pment_dt_ymd_250=payment_date,
                    collected_by_250=collected_by,
                    modified_by_250=request.session['executive_user_id'],  
                    modified_on_250=now(),
                    created_by_250=request.session['executive_user_id'],  
                    created_on_250=now(),
                )

                messages.success(request, "credit added successfully.")
        
                return redirect('add_credit')  
        else:
            form = CreditForm()




        # show table data
        internal_credit_data = Pcash250Addcredit.objects.filter(type_250='internal_credit').order_by('-id_250')

        # Prepare internal_credit_list with bank names
        internal_credit_list = []

        for obj in internal_credit_data:
            try:
                # Retrieve the bank using bank_id_250
                bank = Pcash150Bank.objects.get(id_150=obj.bank_id_250)
                bank_name = bank.bank_full_name_150
            except ObjectDoesNotExist:
                # If no bank found, use a default value (e.g., 'Unknown Bank')
                bank_name = 'Unknown Bank'
            
            # Append data to internal_credit_list with bank name included
            internal_credit_list.append({
                'type_250': 'Internal Credit' if obj.type_250 == 'internal_credit' else '',
                'cash_withdrawal_date': convert_ymd_to_dmy(obj.cash_withdrawal_ymd_250),
                'bank_name': bank_name,  
                'instrument_no_250': obj.instrument_no_250,
                'instrument_type_250': obj.instrument_type_250,
                'instrument_ymd_250': convert_ymd_to_dmy(obj.instrument_ymd_250),
                'amount_250': obj.amount_250,
                'narration_250' : obj.narration_250,
                'withdrawal_by_250': obj.withdrawal_by_250,
            })


        # Prepare other_payment_list similarly
        other_payment_data = Pcash250Addcredit.objects.filter(type_250='other_payment').order_by('-id_250')

        other_payment_list = [
            {
                'type_250': 'Others' if obj.type_250 == 'other_payment' else '',
                'pment_dt_ymd_250': convert_ymd_to_dmy(obj.pment_dt_ymd_250),
                'paid_by_250': obj.paid_by_250,
                'amount_250': obj.amount_250,
                'narration_250' : obj.narration_250,
                'collected_by_250': obj.collected_by_250,
            }
            for obj in other_payment_data
        ]

        # Context
        context = {
            'form': form,
            'internal_credit_list': internal_credit_list,
            'other_payment_list': other_payment_list,
            'status': 'edit',
        }

        return render(request, 'add_credit.html', context)
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def expenses_view(request):
    if 'is_login' in request.session.keys():

        request.session['current_page']='expenses_view'

        if request.method == 'POST':
            form = ExpenseForm(request.POST)
            if form.is_valid():
                # Fetch common data once, defaulting to an empty string if not provided
                executive = form.cleaned_data.get('executive', '')
                expense_date = format_date(form.cleaned_data.get('expense_date', ''))
                total_expense = form.cleaned_data.get('total_expense', '')

                if total_expense <= request.session['cash_in_hand']:
                
                    expn_gist=Pcash300ExpnGist.objects.create(
                        expn_ymd_300=expense_date,
                        user_id_100_300=request.session['executive_user_id'],
                        exec_name_300=executive,
                        total_amount_300=total_expense,
                        modified_by_300=request.session['executive_user_id'],  
                        modified_on_300=now(),
                        created_by_300=request.session['executive_user_id'],  
                        created_on_300=now(),
                    )

                    # Store district-related data in a list for easy iteration
                    district_data = []
                    for i in range(1, 11):  
                        district_data.append({
                            'district': form.cleaned_data.get(f'district_{i}', ''),
                            'narration': form.cleaned_data.get(f'narration_{i}', ''),
                            'amount': form.cleaned_data.get(f'amount_{i}', ''),
                        })

                
                    for data in district_data:
                        if data['district'] and  data['amount']:
                            Pcash350ExpnDetails.objects.create(
                                expn_head_200_350=data['district'],
                                expn_gist_id_350_350=expn_gist.id_300,
                                narration_350=data['narration'],
                                amount_350=data['amount'],
                                modified_by_350=request.session['executive_user_id'],  
                                modified_on_350=now(),
                                created_by_350=request.session['executive_user_id'],  
                                created_on_350=now(),
                            )
                    
                    messages.success(request, "expenses added successfully.")

                    # Redirect after successfully handling the form
                    return redirect('add_expenses')
                else:
                    messages.error(request, "insufficient balance to expense.")



        else:
            form = ExpenseForm()

        return render(request, 'add_expenses.html', {'form': form})
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def dashboard_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'dashboard_view'

        # Fetch the sum of amount_250 from Pcash250Addcredit, converting CharField to FloatField
        total_credit_amount = Pcash250Addcredit.objects.aggregate(
            total=Sum(Cast('amount_250', FloatField()))
        )['total'] or 0.0  # Handle None if no records

        # Fetch the sum of total_amount_300 from Pcash300ExpnGist, converting CharField to FloatField
        total_expense_amount = Pcash300ExpnGist.objects.aggregate(
            total=Sum(Cast('total_amount_300', FloatField()))
        )['total'] or 0.0  # Handle None if no records

        # Fetch the sum of adj_add_amnt_400 and adj_deduct_amnt_400 from Pcash400Adjustment, converting CharField to FloatField
        total_adj_add_amount = Pcash400Adjustment.objects.aggregate(
            total=Sum(Cast('adj_add_amnt_400', FloatField()))
        )['total'] or 0.0  # Handle None if no records

        total_adj_deduct_amount = Pcash400Adjustment.objects.aggregate(
            total=Sum(Cast('adj_deduct_amnt_400', FloatField()))
        )['total'] or 0.0  # Handle None if no records

        # Calculate adjusted cash in hand based on credit, expense, and adjustments
        cash_in_hand = total_credit_amount - total_expense_amount + total_adj_add_amount - total_adj_deduct_amount


        # Round cash in hand to two decimal places
        cash_in_hand = round(cash_in_hand, 2)

        request.session['cash_in_hand'] = cash_in_hand
        

        # Show internal credit data
        internal_credit_data = Pcash250Addcredit.objects.filter(type_250='internal_credit').order_by('-cash_withdrawal_ymd_250')

        # Prepare internal_credit_list with bank names
        internal_credit_list = []
        for obj in internal_credit_data:
            try:
                # Retrieve the bank using bank_id_250
                bank = Pcash150Bank.objects.get(id_150=obj.bank_id_250)
                bank_name = bank.bank_full_name_150
            except ObjectDoesNotExist:
                # If no bank found, use a default value (e.g., 'Unknown Bank')
                bank_name = 'Unknown Bank'

            # Append data to internal_credit_list with bank name included
            internal_credit_list.append({
                'type_250': obj.type_250,
                'cash_withdrawal_date': convert_ymd_to_dmy(obj.cash_withdrawal_ymd_250),
                'bank_name': bank_name,
                'amount_250': obj.amount_250,
            })

        # Context
        context = {
            'internal_credit_list': internal_credit_list,
            'cash_in_hand': cash_in_hand,
        }
        return render(request, 'dashboard.html', context)
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def expense_head_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page']='expense_head_view'

        if request.method == 'POST':
            form = ExpenditureForm(request.POST)
            if form.is_valid():
                # Extract the cleaned form data
                expenditure_head = form.cleaned_data.get('expn_head_name_200', '')
                expense_code = form.cleaned_data.get('expn_code_200', '')
                expn_head_group = form.cleaned_data.get('expn_head_group_200', '')

                Pcash200Expnhead.objects.create(
                            expn_head_group_200=expn_head_group,
                            expn_head_name_200=expenditure_head,
                            expn_code_200=expense_code,
                            active_yn_200='yes',
                            modified_by_200=request.session.get('executive_user_id', 'unknown'),  
                            modified_on_200=now(),
                            created_by_200=request.session.get('executive_user_id', 'unknown'),  
                            created_on_200=now(),
                        )

                
                messages.success(request, "Expenditure Head added successfully.")

                
                return redirect('expend_head')

        else:
            form = ExpenditureForm()

        expend_head_list = Pcash200Expnhead.objects.all().order_by('expn_head_name_200')

        context = {
            'form': form,
            'expend_head_list': expend_head_list,
        }
            
        return render(request, 'expend_head.html', context)
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def edit_expense_head_view(request, pk):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'edit_expense_head_view'

        # Fetch the specific record to be edited
        expense_head = get_object_or_404(Pcash200Expnhead, id_200=pk)

        if request.method == 'POST':
            form = ExpenditureForm(request.POST, instance=expense_head)
            if form.is_valid():
                # Extract the cleaned form data
                expenditure_head = form.cleaned_data.get('expn_head_name_200', '')
                expense_code = form.cleaned_data.get('expn_code_200', '')
                expn_head_group = form.cleaned_data.get('expn_head_group_200', '')

                # Update the record
                expense_head.expn_head_group_200 = expn_head_group
                expense_head.expn_head_name_200 = expenditure_head
                expense_head.expn_code_200 = expense_code
                expense_head.active_yn_200 = 'yes'
                expense_head.modified_by_200 = request.session.get('executive_user_id', 'unknown')
                expense_head.modified_on_200 = now()
                expense_head.save()

                messages.success(request, "Expenditure Head updated successfully.")
                return redirect('expend_head')

        else:
            form = ExpenditureForm(instance=expense_head)

        expend_head_list = Pcash200Expnhead.objects.all().order_by('expn_head_name_200')

        context = {
            'form': form,
            'expend_head_list': expend_head_list,
            'status':'edit',
        }
        
        return render(request, 'expend_head.html', context)
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def cancel_expense_head_view(request, pk):
    if 'is_login' in request.session.keys():
        expense_head = get_object_or_404(Pcash200Expnhead, id_200=pk)
        expense_head.delete()
        messages.success(request, "Expenditure Head deleted successfully.")
        return redirect('expend_head')
    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def adjustment_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'adjustment_view'
        request.session['current_value'] = '1'

        if request.method == 'POST':
            adjustment_type = request.POST['adjustment_type']

            if adjustment_type == 'credit_adjust':
                request.session['adj_type'] = adjustment_type
                return redirect('check_validation')

            if adjustment_type == 'expend_adjust':
                request.session['adj_type'] = adjustment_type
                return redirect('check_validation')

        
        adjustments = Pcash400Adjustment.objects.all().order_by('-id_400')

        return render(request, 'adjust_entry_1.html', {'adjustments': adjustments})

    else:
        return redirect('login')

##########################################################################################################################################

@never_cache
def check_validation(request):
    request.session['current_value'] = '2'
    if request.method == 'POST':
            form = CreditValidationForm(request.POST)
            if form.is_valid():
                adjustment_type = request.session['adj_type']
                credit_type = request.POST['credit_type']

                request.session['credit_type'] = credit_type

                if adjustment_type == 'credit_adjust':

                    if credit_type == 'internal_credit':
                        # Retrieving form data
                        cash_withdrawal_date = format_date(form.cleaned_data.get('withdrawal_date'))  
                        instrument_type = form.cleaned_data.get('instrument_type') 
                        cheque_number = form.cleaned_data.get('cheque_number') 

                        request.session['cash_withdrawal_date'] = cash_withdrawal_date
                        request.session['instrument_type'] = instrument_type
                        request.session['cheque_number'] = cheque_number

                        credit_data = Pcash250Addcredit.objects.filter(type_250=credit_type, instrument_no_250=cheque_number, cash_withdrawal_ymd_250=cash_withdrawal_date, instrument_type_250=instrument_type).first()  
                        
                        if credit_data:
                            request.session['source_type_400'] = credit_data.type_250
                            request.session['select_trnxn_dt_ymd_400'] = credit_data.cash_withdrawal_ymd_250
                            request.session['amount_to_be_adjted_400'] = credit_data.amount_250
                            
                            return redirect('adjust_view')  
                        else:
                            messages.error(request, "No data found.")
                            

                    if credit_type == 'other_payment':
                        # Retrieving form data 
                        payment_date = format_date(form.cleaned_data.get('payment_date'))  
                        paid_by = form.cleaned_data.get('paid_by')  

                        credit_data = Pcash250Addcredit.objects.filter(type_250=credit_type, pment_dt_ymd_250=payment_date, paid_by_250 = paid_by).first()  

                        if credit_data:
                            request.session['payment_date'] = payment_date
                            request.session['paid_by'] = paid_by
                            request.session['source_type_400'] = credit_data.type_250
                            request.session['select_trnxn_dt_ymd_400'] = credit_data.pment_dt_ymd_250
                            request.session['amount_to_be_adjted_400'] = credit_data.amount_250
                            return redirect('adjust_view')  
                        else:
                            messages.error(request, "No data found.") 

                if adjustment_type ==  'expend_adjust':
                    expense_date = format_date(form.cleaned_data.get('expense_date'))  
                    total_amount = form.cleaned_data.get('executive') 

                    epense_data = Pcash300ExpnGist.objects.filter(total_amount_300=total_amount, expn_ymd_300 = expense_date).first()  

                    if epense_data:
                        request.session['id_300'] = epense_data.id_300
                        request.session['total_amount_300'] = epense_data.total_amount_300
                        request.session['expn_ymd_300'] = epense_data.expn_ymd_300
                        request.session['value'] = 'expense'
                        return redirect('adjust_view')  
                    else:
                        messages.error(request, "No data found.")  
    else:
        form = CreditValidationForm()

    adjustments = Pcash400Adjustment.objects.all().order_by('-id_400')

    return render(request, 'adjust_entry_1.html',{'form': form,'adjustments': adjustments})

##########################################################################################################################################

@never_cache
def adjust_view(request):
    request.session['current_value'] = '3'

    def get_initial_data():
        """ Helper function to retrieve initial data based on adjustment type """
        if request.session.get('adj_type') == 'credit_adjust':
            if request.session.get('credit_type') == 'internal_credit':
                return {
                    'withdrawal_date': datetime.strptime(request.session.get('cash_withdrawal_date'), "%Y%m%d").strftime("%d/%m/%Y"),
                    'instrument_type': request.session.get('instrument_type'),
                    'cheque_number': request.session.get('cheque_number')
                }
            elif request.session.get('credit_type') == 'other_payment':
                return {
                    'payment_date': datetime.strptime(request.session.get('payment_date'), "%Y%m%d").strftime("%d/%m/%Y"),
                    'paid_by': request.session.get('paid_by'),
                }
        elif request.session.get('adj_type') == 'expend_adjust':
            return {
                'expense_date': datetime.strptime(request.session.get('expn_ymd_300'), "%Y%m%d").strftime("%d/%m/%Y"),
                'executive': request.session.get('total_amount_300'),
            }
        return {}

    if request.method == 'POST':
        form_2 = AdjustmentForm(request.POST)
        if form_2.is_valid():
            adjustment_type = request.session['adj_type']
            amount = form_2.cleaned_data.get('amount')
            operation = form_2.cleaned_data.get('operation')
            narration = form_2.cleaned_data.get('narration')
            obj = Pcash400Adjustment()

            def handle_invalid_amount():
                form_2.add_error('amount', "Enter a valid amount.")
                form = CreditValidationForm(initial=get_initial_data())
                adjustments = Pcash400Adjustment.objects.all().order_by('-id_400')
                return render(request, 'adjust_entry_1.html', {'form': form, 'form_2': form_2, 'adjustments': adjustments})

            if adjustment_type == 'credit_adjust':
                if operation == 'minus' and (float(amount) >= float(request.session.get('amount_to_be_adjted_400')) or float(amount) >= request.session['cash_in_hand']):
                    return handle_invalid_amount()
            
            
            if adjustment_type == 'expend_adjust':
                if operation == 'minus' and (float(amount) >= float(request.session.get('total_amount_300')) or float(amount) >= request.session['cash_in_hand']):
                    return handle_invalid_amount()
                

            if operation == 'plus':
                obj.adj_add_amnt_400 = amount
            elif operation == 'minus':
                obj.adj_deduct_amnt_400 = amount

            # Common data assignment
            obj.adj_ymd_400 = get_current_date_ymd()
            obj.adj_narration_400 = narration
            obj.modified_by_400 = request.session['executive_user_id']
            obj.modified_on_400 = now()
            obj.created_by_400 = request.session['executive_user_id']
            obj.created_on_400 = now()

            if adjustment_type == 'credit_adjust':

                obj.adj_type_400 = request.session.get('adj_type')
                obj.source_type_400 = request.session.get('source_type_400')
                obj.select_trnxn_dt_ymd_400 = request.session.get('select_trnxn_dt_ymd_400')
                obj.amount_to_be_adjted_400 = request.session.get('amount_to_be_adjted_400')
                obj.expn_details_id_350_400 = 0

                messages.success(request, "Credit adjustment successfully.")
            elif adjustment_type == 'expend_adjust':

                obj.adj_type_400 = 'expn_adj'
                obj.source_type_400 = 'expn'
                obj.select_trnxn_dt_ymd_400 = request.session.get('expn_ymd_300')
                obj.amount_to_be_adjted_400 = request.session.get('total_amount_300')
                obj.expn_details_id_350_400 = request.session.get('id_300')

                messages.success(request, "Expenditure adjustment successfully.")

            obj.save()

            request.session['current_value'] = '1'
            return redirect('adjustment')

    else:
        initial_data = get_initial_data()

    form = CreditValidationForm(initial=initial_data)
    form_2 = AdjustmentForm()
    adjustments = Pcash400Adjustment.objects.all().order_by('-id_400')

    return render(request, 'adjust_entry_1.html', {'form': form, 'form_2': form_2, 'adjustments': adjustments})

##########################################################################################################################################
'''
class MisData:
    def __init__(self):
        self.transaction_date = None
        self.reference_no = None
        self.details = None
        self.credit_amount = None
        self.debit_amount = None
        self.total_amount = None
        self.expn_gist = None
        self.bank_id = None
        self.withdrawl_by = None
        self.paid_by = None


@never_cache
def mis_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'mis_view'
        
        if request.method == 'POST':
            start_date = request.POST.get('payment_date', None)
            end_date = request.POST.get('withdrawal_date', None)
            
            if start_date and end_date:
                try:
                    # Parse input dates in DD/MM/YYYY format
                    start_date_obj = datetime.strptime(start_date, "%d/%m/%Y")
                    end_date_obj = datetime.strptime(end_date, "%d/%m/%Y")

                    # Format as YYYYMMDD (assuming this is how dates are stored in the database)
                    start_formatted_date = start_date_obj.strftime("%Y%m%d")
                    end_formatted_date = end_date_obj.strftime("%Y%m%d")

                    request.session['start_formatted_date'] = start_formatted_date
                    request.session['end_formatted_date'] = end_formatted_date

                    # Query Pcash250Addcredit model for the date range
                    addcredit_results = Pcash250Addcredit.objects.filter(
                        cash_withdrawal_ymd_250__range=[start_formatted_date, end_formatted_date]
                    )

                    # Query Pcash300ExpnGist model for the date range
                    expngist_results = Pcash300ExpnGist.objects.filter(
                        expn_ymd_300__range=[start_formatted_date, end_formatted_date]
                    )

                    combined_results = []

                    for i in range(0,len(list(addcredit_results))):
                        obj = MisData()
                        obj.transaction_date = addcredit_results[i].cash_withdrawal_ymd_250
                        obj.details = addcredit_results[i].type_250
                        obj.credit_amount , obj.total_amount = addcredit_results[i].amount_250, addcredit_results[i].amount_250

                        if obj.details == 'internal_credit':
                            obj.withdrawl_by = addcredit_results[i].withdrawal_by_250
                            obj.bank_id = addcredit_results[i].bank_id_250
                            obj.reference_no = addcredit_results[i].instrument_no_250
                        
                        if obj.details == 'other_payment':
                            obj.paid_by = addcredit_results[i].paid_by_250

                        combined_results.append(obj)

                    for i in range(0,len(list(expngist_results))):
                        obj = MisData()
                        obj.transaction_date = expngist_results[i].expn_ymd_300
                        obj.details = 'Expense_Head'
                        obj.total_amount = expngist_results[i].total_amount_300
                        obj.expn_gist = expngist_results[i].id_300

                        
                        combined_results.append(obj)
                
                    request.session['total_value'] = 0.0
                    combined_results.sort(key=lambda x: x.transaction_date)
                    return render(request, 'mis.html', {'results': combined_results})

                except ValueError:
                    # Handle invalid date format error
                    messages.error(request, "Invalid date format. Use DD/MM/YYYY.")
            else:
                # Handle missing dates
                messages.error(request, "Please provide both start and end dates.")
            
        return render(request, 'mis.html', {'results': None})
    else:
        return redirect('login')
'''
#########################################################################################################################################


@never_cache
def another_mis_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'another_mis'
        
        if request.method == 'POST':
            start_date = request.POST.get('payment_date', None)
            end_date = request.POST.get('withdrawal_date', None)
            district = request.POST.get('district', None)

            request.session['district'] = district

            
            if start_date and end_date:
                try:
                    # Parse input dates in DD/MM/YYYY format
                    start_date_obj = datetime.strptime(start_date, "%d/%m/%Y")
                    end_date_obj = datetime.strptime(end_date, "%d/%m/%Y")

                    if start_date_obj > end_date_obj:
                        messages.error(request, "Start date should be less than or equal to an end date.")

                    # Format as YYYYMMDD (assuming this is how dates are stored in the database)
                    start_formatted_date = start_date_obj.strftime("%Y%m%d")
                    end_formatted_date = end_date_obj.strftime("%Y%m%d")


                    if district == 'Credit':

                        # Query Pcash250Addcredit model for the date range
                        addcredit_results = Pcash250Addcredit.objects.filter(
                            cash_withdrawal_ymd_250__range=[start_formatted_date, end_formatted_date]
                        )

                        return render(request, 'mis_2.html', {'results': addcredit_results})
                    
                    
                    if district == 'Expense':

                        # Query Pcash300ExpnGist model for the date range
                        expngist_results = Pcash300ExpnGist.objects.filter(
                            expn_ymd_300__range=[start_formatted_date, end_formatted_date]
                        )

                        return render(request, 'mis_2.html', {'results': expngist_results})

                    
                except ValueError:
                    # Handle invalid date format error
                    messages.error(request, "Invalid date format. Use DD/MM/YYYY.")
            else:
                # Handle missing dates
                messages.error(request, "Please provide both start and end dates.")
            
        return render(request, 'mis_2.html', {'results': None})
    else:
        return redirect('login')

#########################################################################################################################################






class MisData:
    def __init__(self):
        self.transaction_date = None
        self.reference_no = None
        self.details = None
        self.credit_amount = None
        self.debit_amount = None
        self.total_amount = None
        self.expn_gist = None
        self.bank_id = None
        self.withdrawl_by = None
        self.paid_by = None
        self.credit_adjustment = None
        self.expense_adjustment = None
        self.adjustment_type = None
        self.balance = None


@never_cache
def mis_view(request):
    if 'is_login' in request.session.keys():
        request.session['current_page'] = 'mis_view'
        
        if request.method == 'POST':
            start_date = request.POST.get('payment_date', None)
            end_date = request.POST.get('withdrawal_date', None)
            
            if start_date and end_date:
                try:
                    # Parse input dates in DD/MM/YYYY format
                    start_date_obj = datetime.strptime(start_date, "%d/%m/%Y")
                    end_date_obj = datetime.strptime(end_date, "%d/%m/%Y")

                    # Format as YYYYMMDD (assuming this is how dates are stored in the database)
                    start_formatted_date = start_date_obj.strftime("%Y%m%d")
                    end_formatted_date = end_date_obj.strftime("%Y%m%d")

                    request.session['start_formatted_date'] = start_formatted_date
                    request.session['end_formatted_date'] = end_formatted_date

                    # # Query Pcash250Addcredit model for the date range
                    # addcredit_results = Pcash250Addcredit.objects.filter(
                    #     cash_withdrawal_ymd_250__range=[start_formatted_date, end_formatted_date]
                    # )

                    # # Query Pcash300ExpnGist model for the date range
                    # expngist_results = Pcash300ExpnGist.objects.filter(
                    #     expn_ymd_300__range=[start_formatted_date, end_formatted_date]
                    # )

                    # # Query Pcash400Adjustment model for the date range
                    # adjustment_results = Pcash400Adjustment.objects.filter(
                    #     adj_ymd_400__range=[start_formatted_date, end_formatted_date]
                    # )



                    # Query Pcash250Addcredit model for the date range
                    addcredit_results = Pcash250Addcredit.objects.all()
                    # Query Pcash300ExpnGist model for the date range
                    expngist_results = Pcash300ExpnGist.objects.all()

                    # Query Pcash400Adjustment model for the date range
                    adjustment_results = Pcash400Adjustment.objects.all()

                    combined_results = []

                    for i in range(0,len(list(addcredit_results))):
                        obj = MisData()
                        obj.transaction_date = addcredit_results[i].cash_withdrawal_ymd_250
                        obj.details = addcredit_results[i].type_250
                        obj.credit_amount , obj.total_amount = addcredit_results[i].amount_250, addcredit_results[i].amount_250

                        if obj.details == 'internal_credit':
                            obj.withdrawl_by = addcredit_results[i].withdrawal_by_250
                            obj.bank_id = addcredit_results[i].bank_id_250
                            obj.reference_no = addcredit_results[i].instrument_no_250
                        
                        if obj.details == 'other_payment':
                            obj.paid_by = addcredit_results[i].paid_by_250

                        combined_results.append(obj)

                    for i in range(0,len(list(expngist_results))):
                        obj = MisData()
                        obj.transaction_date = expngist_results[i].expn_ymd_300
                        obj.details = 'Expense_Head'
                        obj.total_amount = expngist_results[i].total_amount_300
                        obj.expn_gist = expngist_results[i].id_300

                        
                        combined_results.append(obj)

                    for i in range(0,len(list(adjustment_results))):
                        obj = MisData()
                        obj.transaction_date = adjustment_results[i].adj_ymd_400
                        obj.details = 'adjustment'
                        obj.adjustment_type = adjustment_results[i].source_type_400

                        if adjustment_results[i].adj_add_amnt_400:
                            obj.credit_adjustment = adjustment_results[i].adj_add_amnt_400
                        else:
                            obj.expense_adjustment = adjustment_results[i].adj_deduct_amnt_400


                        combined_results.append(obj)
                        
                        
                
                    combined_results.sort(key=lambda x: x.transaction_date)


                    
                    total_value = 0.0

                    for obj in combined_results:
                        if obj.details == 'internal_credit' or obj.details == 'other_payment':
                            total_value = total_value + float(obj.total_amount) 
                        elif obj.details == 'Expense_Head':
                            total_value = total_value - float(obj.total_amount) 
                        elif obj.details == 'adjustment':
                            if  obj.credit_adjustment:
                                total_value = total_value + float(obj.credit_adjustment)
                            else:
                                total_value = total_value - float(obj.expense_adjustment)

                        obj.balance = total_value

                    filtered_results = [
                        obj for obj in combined_results
                        if start_formatted_date <= obj.transaction_date <= end_formatted_date
                    ]
                                        

                    return render(request, 'mis.html', {'results': filtered_results})

                except ValueError:
                    # Handle invalid date format error
                    messages.error(request, "Invalid date format. Use DD/MM/YYYY.")
            else:
                # Handle missing dates
                messages.error(request, "Please provide both start and end dates.")
            
        return render(request, 'mis.html', {'results': None})
    else:
        return redirect('login')

#########################################################################################################################################















@never_cache
def export_mis_data_to_excel(request):
    if 'is_login' in request.session.keys():
        # Create a new workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = 'MIS Data'

        # Define headers and their styles
        headers = [
            'Transaction Date', 'Reference Number', 'Details', 'Credit Amount', 'Debit Amount', 'Total', 'Balance'
        ]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0072C6", end_color="0072C6", fill_type="solid")
        header_border = Border(bottom=Side(style="thin"))

        # Write headers with styling
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
            cell.border = header_border

        start_formatted_date = request.session.get('start_formatted_date')
        end_formatted_date = request.session.get('end_formatted_date')

        # Query models for the date range
        addcredit_results = Pcash250Addcredit.objects.filter(
            cash_withdrawal_ymd_250__range=[start_formatted_date, end_formatted_date]
        )
        expngist_results = Pcash300ExpnGist.objects.filter(
            expn_ymd_300__range=[start_formatted_date, end_formatted_date]
        )

        combined_results = []

        # Process addcredit_results
        for result in addcredit_results:
            obj = {
                'transaction_date': datetime.strptime(result.cash_withdrawal_ymd_250, '%Y%m%d').strftime('%d/%m/%Y'),
                'details': result.type_250,
                'credit_amount': float(result.amount_250),
                'total_amount': float(result.amount_250),
                'reference_no': result.instrument_no_250,
                'withdrawl_by': result.withdrawal_by_250,
                'bank_id': result.bank_id_250,
                'paid_by': result.paid_by_250,
            }

            if obj['details'] == 'internal_credit':
                try:
                    bank = Pcash150Bank.objects.get(id_150=obj['bank_id'])
                    bank_full_name = bank.bank_full_name_150 if bank.bank_full_name_150 else "Unknown Bank"
                    details_text = f"Internal Credit: {bank_full_name} - Withdrawn by: {obj['withdrawl_by']}"
                    obj['details'] = details_text
                except Pcash150Bank.DoesNotExist:
                    obj['details'] = f"Internal Credit: Unknown Bank - Withdrawn by: {obj['withdrawl_by']}"

            elif obj['details'] == 'other_payment':
                obj['details'] = f"Other Payment: Paid by: {obj['paid_by']}"

            combined_results.append(obj)

        # Process expngist_results
        for result in expngist_results:
            expn_details = Pcash350ExpnDetails.objects.filter(expn_gist_id_350_350=result.id_300)

            for detail in expn_details:
                try:
                    expn_head = Pcash200Expnhead.objects.get(id_200=detail.expn_head_200_350)
                    expn_head_name = expn_head.expn_head_name_200
                except Pcash200Expnhead.DoesNotExist:
                    expn_head_name = "Unknown"

                obj = {
                    'transaction_date': datetime.strptime(result.expn_ymd_300, '%Y%m%d').strftime('%d/%m/%Y'),
                    'details': f"Expense Head: {expn_head_name}",
                    'total_amount': float(result.total_amount_300),
                    'amount': float(detail.amount_350) if detail.amount_350 else 0.0,
                    'credit_amount': 0.0,  # No credit for expenses
                    'reference_no': '',  # No reference number for expenses
                }
                combined_results.append(obj)

        # Sort combined results by transaction date
        combined_results.sort(key=lambda x: x['transaction_date'])

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                             top=Side(style='thin'), bottom=Side(style='thin'))
        current_balance = 0.0

        last_transaction_date = None
        last_total_amount = None

        for row_num, transaction in enumerate(combined_results, start=2):
            # Write the transaction date only if it's different from the previous one
            if transaction['details'].startswith('Expense Head'):
                if transaction['transaction_date'] != last_transaction_date:
                    ws.cell(row=row_num, column=1, value=transaction['transaction_date']).border = thin_border
                    last_transaction_date = transaction['transaction_date']
                else:
                    ws.cell(row=row_num, column=1, value='').border = thin_border  # Empty for repeated dates

            else:
                ws.cell(row=row_num, column=1, value=transaction['transaction_date']).border = thin_border

            # Reference Number
            if transaction['details'].startswith('Internal Credit'):
                ws.cell(row=row_num, column=2, value=transaction.get('reference_no', '')).border = thin_border
            else:
                ws.cell(row=row_num, column=2, value='').border = thin_border

            # Details
            details_text = transaction['details']
            ws.cell(row=row_num, column=3, value=details_text).border = thin_border

            # Credit Amount
            credit_amount = transaction.get('credit_amount', 0)
            if transaction['details'].startswith('Expense Head'):
                ws.cell(row=row_num, column=4, value='').border = thin_border  # Empty for expenses
            else:
                ws.cell(row=row_num, column=4, value=credit_amount).border = thin_border

            # Debit Amount
            debit_amount = transaction.get('amount', 0) if 'amount' in transaction else ''
            ws.cell(row=row_num, column=5, value=debit_amount).border = thin_border

            # Total Amount
            total_amount = transaction.get('total_amount', 0)
            if transaction['details'].startswith('Expense Head'):
                if total_amount != last_total_amount:
                    ws.cell(row=row_num, column=6, value=total_amount).border = thin_border
                    last_total_amount = total_amount
                else:
                    ws.cell(row=row_num, column=6, value='').border = thin_border  # Empty for repeated total amounts
            else:
                ws.cell(row=row_num, column=6, value=total_amount).border = thin_border

            # Calculate new balance
            if credit_amount:
                current_balance += total_amount  # For credit transactions
            elif debit_amount:
                current_balance -= debit_amount  # For debit transactions
            
            # Balance
            ws.cell(row=row_num, column=7, value=current_balance).border = thin_border

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column].width = adjusted_width

        # Save workbook to a BytesIO object
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Prepare response for download
        filename = "mis.xlsx"
        response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        return response
    else:
        return redirect('/')

#########################################################################################################################################

@never_cache
def search_view(request):
    if 'is_login' in request.session.keys():
        if request.method == 'POST':
            search_data = request.POST.get('search_data', '')
            credit_type = request.POST.get('credit_type', '')

            request.session['credit_type'] = credit_type

            # Initialize lists to hold results
            internal_credit_list = []
            other_payment_list = []

            if search_data:  # Only search if there's input
                if credit_type == 'internal_credit':
                    internal_credit_data = Pcash250Addcredit.objects.filter(
                        Q(type_250__icontains=search_data) |
                        Q(paid_by_250__icontains=search_data) |
                        Q(bank_id_250__icontains=search_data) |
                        Q(instrument_type_250__icontains=search_data) |
                        Q(instrument_no_250__icontains=search_data) |
                        Q(amount_250__icontains=search_data) |
                        Q(withdrawal_by_250__icontains=search_data)
                    )

                    for obj in internal_credit_data:
                        internal_credit_list.append({
                            'type_250': 'Internal Credit' if obj.type_250 == 'internal_credit' else '',
                            'cash_withdrawal_date': convert_ymd_to_dmy(obj.cash_withdrawal_ymd_250),
                            'instrument_no_250': obj.instrument_no_250,
                            'instrument_type_250': obj.instrument_type_250,
                            'instrument_ymd_250': convert_ymd_to_dmy(obj.instrument_ymd_250),
                            'amount_250': obj.amount_250,
                            'narration_250': obj.narration_250,
                            'withdrawal_by_250': obj.withdrawal_by_250,
                        })

                elif credit_type == 'other_payment':
                    other_payment_data = Pcash250Addcredit.objects.filter(
                        Q(type_250__icontains=search_data) |
                        Q(amount_250__icontains=search_data) |
                        Q(pment_dt_ymd_250__icontains=search_data) |
                        Q(narration_250__icontains=search_data) |
                        Q(collected_by_250__icontains=search_data)
                    )

                    other_payment_list = [
                        {
                            'type_250': 'Others' if obj.type_250 == 'other_payment' else '',
                            'pment_dt_ymd_250': convert_ymd_to_dmy(obj.pment_dt_ymd_250),
                            'paid_by_250': obj.paid_by_250,
                            'amount_250': obj.amount_250,
                            'narration_250': obj.narration_250,
                            'collected_by_250': obj.collected_by_250,
                        }
                        for obj in other_payment_data
                    ]

            # Context
            context = {
                'internal_credit_list': internal_credit_list,
                'other_payment_list': other_payment_list,
                'status': 'edit',
                'no_results': not internal_credit_list and not other_payment_list,
            }

            return render(request, 'search_data.html', context)

    else:
        return redirect('login')

##########################################################################################################################################
